"""
Pittsburgh Winter Storm Salt Truck Route Optimization - Production Version
Integrates OpenStreetMap, real-time weather, and traffic data
"""
 
import osmnx as ox
import networkx as nx
import numpy as np
import requests
from dataclasses import dataclass, field
from typing import List, Dict, Set, Tuple, Optional
from collections import defaultdict
import heapq
from datetime import datetime
import json
import folium
from scipy.spatial import KDTree

#Configuration
PITTSBURGH_COORDS = (40.4406, -79.9959)
PITTSBURGH_BBOX = (40.361, 40.501, -80.095, -79.865)  # South, North, West, East

#============================================================================
# JURISDICTION FILTERING - CITY OF PITTSBURGH ROADS ONLY
#============================================================================
#Excludes roads maintained by PennDOT (state) and Allegheny County
#City of Pittsburgh DPW is responsible for ~900 centerline miles

CITY_ONLY_MODE = True  #Set to False to include all roads (city + county + state)

#Exclude these highway types (maintained by PennDOT)
EXCLUDE_HIGHWAY_TYPES = [
    'motorway',       #Interstate highways (I-376, I-279, I-579)
    'motorway_link',  #Highway on/off ramps
    'trunk',          #Major US routes (US-19, US-22, US-30)
]

#Explicit list of roads maintained by PennDOT (not city responsibility)
PENNDOT_MAINTAINED_ROADS = {
    #Interstate highways
    'Interstate 376', 'I-376', 'I 376', 'Parkway East', 'Parkway West',
    'Interstate 279', 'I-279', 'I 279', 'Parkway North',
    'Interstate 579', 'I-579', 'I 579', 'Crosstown Boulevard',
    
    #US Routes
    'McKnight Road', 'US 19', 'Route 19', 'US-19',
    'Penn Lincoln Parkway', 'US 22', 'US-22', 'US 30', 'US-30',
    
    #State Routes
    'PA 28', 'Route 28', 'PA-28',
    'PA 51', 'Route 51', 'PA-51', 'Saw Mill Run Boulevard',
    'PA 837', 'Route 837', 'PA-837',
    'PA 885', 'Route 885', 'PA-885',
    
    #Major PennDOT-maintained bridges
    'Fort Pitt Bridge',
    'Fort Duquesne Bridge', 
    'Veterans Bridge',
    'West End Bridge',
}

@dataclass
class WeatherConditions:
    """Current weather conditions affecting road treatment"""
    temperature_f: float
    precipitation_type: str  # snow, sleet, freezing_rain, none
    precipitation_rate: float  # inches per hour
    wind_speed_mph: float
    visibility_miles: float
    road_temperature_f: float
    timestamp: datetime
    
    def get_severity_multiplier(self) -> float:
        """Calculate how much weather impacts salting needs"""
        multiplier = 1.0
        
        #Temperature impact
        if self.temperature_f < 15:
            multiplier *= 1.5  # Very cold, salt less effective
        elif self.temperature_f < 25:
            multiplier *= 1.2
        
        #Precipitation impact
        if self.precipitation_type == "freezing_rain":
            multiplier *= 1.8  # Most dangerous
        elif self.precipitation_type == "sleet":
            multiplier *= 1.5
        elif self.precipitation_type == "snow" and self.precipitation_rate > 1.0:
            multiplier *= 1.4  # Heavy snow
        elif self.precipitation_type == "snow":
            multiplier *= 1.2
        
        #Wind impact (blowing snow)
        if self.wind_speed_mph > 25:
            multiplier *= 1.2
        
        return multiplier

@dataclass
class TrafficData:
    """Real-time traffic conditions"""
    road_id: str
    current_speed_mph: float
    free_flow_speed_mph: float
    congestion_level: float  # 0-1, where 1 is completely congested
    incident_nearby: bool
    timestamp: datetime
    
    def get_priority_boost(self) -> float:
        """Higher congestion = higher priority to clear"""
        if self.incident_nearby:
            return -100  # Significant priority boost
        return -50 * self.congestion_level

@dataclass
class EnhancedRoad:
    """Enhanced road with OSM data and classifications"""
    osm_id: str
    name: str
    highway_type: str  # motorway, trunk, primary, secondary, residential, etc.
    start_node: int
    end_node: int
    length_meters: float
    geometry: List[Tuple[float, float]]  # lat, lon coordinates
    lanes: int
    maxspeed_mph: Optional[float]
    surface: Optional[str]
    priority: int
    avg_daily_traffic: int = 0
    is_bridge: bool = False
    is_tunnel: bool = False
    grade_percent: float = 0.0  # Slope/steepness
    connects_hospital: bool = False
    connects_fire_station: bool = False
    connects_school: bool = False
    is_bus_route: bool = False
    
    def salt_time_minutes(self, weather: WeatherConditions) -> float:
        """Estimate time to salt based on road characteristics and weather"""
        base_time_per_mile = 3.0  # minutes per mile
        length_miles = self.length_meters / 1609.34
        
        base_time = length_miles * base_time_per_mile * self.lanes
        
        #Weather adjustment
        base_time *= weather.get_severity_multiplier()
        
        #Road characteristic adjustments
        if abs(self.grade_percent) > 5:
            base_time *= 1.4  # Steep hills take longer
        if self.is_bridge:
            base_time *= 1.3  # Bridges ice first, need more attention
        if self.is_tunnel:
            base_time *= 0.9  # Tunnels often need less
        
        return base_time

@dataclass
class Depot:
    """Salt truck depot/garage"""
    id: str
    name: str
    latitude: float
    longitude: float
    available_trucks: int
    salt_capacity_tons: float
    osm_node: Optional[int] = None
    refill_available: bool = True
    refill_time_minutes: float = 18.0  # Average time to reload salt
    avg_refill_queue_minutes: float = 5.0  # Wait time if multiple trucks

@dataclass
class Truck:
    """Enhanced truck model with capacity and operational constraints"""
    id: int
    depot_id: str
    tank_capacity_tons: float = 12.0  # Typical salt truck capacity
    current_salt_tons: float = 12.0  # Start full
    fuel_capacity_gallons: float = 100.0
    current_fuel_gallons: float = 100.0
    mpg_spreading: float = 5.0  # MPG while spreading salt
    mpg_traveling: float = 8.0  # MPG while traveling empty
    application_rate_lbs_per_lane_mile: float = 300.0  # Standard rate
    driver_shift_hours: float = 8.0  # Standard shift length
    current_shift_hours: float = 0.0
    break_required_after_hours: float = 4.0
    last_break_hours_ago: float = 0.0
    
    def salt_remaining_lbs(self) -> float:
        """Get remaining salt in pounds"""
        return self.current_salt_tons * 2000
    
    def salt_needed_for_road(self, road: EnhancedRoad) -> float:
        """Calculate salt needed for a road segment in pounds"""
        length_miles = road.length_meters / 1609.34
        return length_miles * road.lanes * self.application_rate_lbs_per_lane_mile
    
    def can_service_road(self, road: EnhancedRoad) -> bool:
        """Check if truck has enough salt for this road"""
        return self.salt_needed_for_road(road) <= self.salt_remaining_lbs()
    
    def miles_until_salt_empty(self) -> float:
        """Calculate miles truck can travel before running out of salt"""
        if self.current_salt_tons <= 0:
            return 0
        lbs_remaining = self.salt_remaining_lbs()
        return lbs_remaining / self.application_rate_lbs_per_lane_mile
    
    def needs_refill(self, upcoming_miles: float, lanes: int = 2) -> bool:
        """Check if refill needed before servicing upcoming route"""
        salt_needed = upcoming_miles * lanes * self.application_rate_lbs_per_lane_mile
        return salt_needed > self.salt_remaining_lbs()
    
    def refill_salt(self, refill_time_minutes: float) -> float:
        """Refill salt tank and return time taken"""
        self.current_salt_tons = self.tank_capacity_tons
        return refill_time_minutes
    
    def use_salt(self, pounds: float):
        """Deduct salt used"""
        tons_used = pounds / 2000
        self.current_salt_tons = max(0, self.current_salt_tons - tons_used)
    
    def needs_break(self) -> bool:
        """Check if driver needs mandatory break"""
        return self.last_break_hours_ago >= self.break_required_after_hours
    
    def take_break(self, break_minutes: float = 30.0):
        """Take mandatory break"""
        self.current_shift_hours += break_minutes / 60
        self.last_break_hours_ago = 0.0
    
    def can_continue_shift(self, additional_hours: float = 0) -> bool:
        """Check if driver can continue working"""
        return (self.current_shift_hours + additional_hours) <= self.driver_shift_hours
    
    def add_work_time(self, minutes: float):
        """Add time to current shift"""
        hours = minutes / 60
        self.current_shift_hours += hours
        self.last_break_hours_ago += hours

@dataclass
class RouteSegment:
    """Represents a segment in a truck's route (road, refill, or break)"""
    segment_type: str  # 'road', 'refill', 'break', 'travel'
    description: str
    time_minutes: float
    distance_miles: float = 0.0
    salt_used_lbs: float = 0.0
    geometry: List[Tuple[float, float]] = field(default_factory=list)
    road_data: Optional[Dict] = None

class WeatherAPI:
    """Fetch real-time weather data"""
    
    @staticmethod
    def get_current_conditions(lat: float, lon: float) -> WeatherConditions:
        """
        Get current weather from National Weather Service API (free, no key needed)
        """
        try:
            #Get grid point
            point_url = f"https://api.weather.gov/points/{lat},{lon}"
            response = requests.get(point_url, headers={'User-Agent': 'PittsburghSaltRouter/1.0'})
            
            if response.status_code == 200:
                data = response.json()
                forecast_url = data['properties']['forecastHourly']
                
                #Get hourly forecast
                forecast_response = requests.get(forecast_url, headers={'User-Agent': 'PittsburghSaltRouter/1.0'})
                
                if forecast_response.status_code == 200:
                    forecast_data = forecast_response.json()
                    current = forecast_data['properties']['periods'][0]
                    
                    temp_f = current['temperature']
                    
                    #Parse precipitation and conditions
                    precip_type = "none"
                    precip_rate = 0.0
                    short_forecast = current['shortForecast'].lower()
                    
                    if "freezing rain" in short_forecast or "ice" in short_forecast:
                        precip_type = "freezing_rain"
                        precip_rate = 0.3
                    elif "sleet" in short_forecast:
                        precip_type = "sleet"
                        precip_rate = 0.2
                    elif "snow" in short_forecast:
                        precip_type = "snow"
                        if "heavy" in short_forecast:
                            precip_rate = 1.5
                        else:
                            precip_rate = 0.5
                    
                    wind_speed = float(current['windSpeed'].split()[0]) if current['windSpeed'] else 10.0
                    
                    #Estimate road temperature (typically 2-5°F colder than air temp)
                    road_temp = temp_f - 3
                    
                    return WeatherConditions(
                        temperature_f=temp_f,
                        precipitation_type=precip_type,
                        precipitation_rate=precip_rate,
                        wind_speed_mph=wind_speed,
                        visibility_miles=10.0,
                        road_temperature_f=road_temp,
                        timestamp=datetime.now()
                    )
        
        except Exception as e:
            print(f"Weather API error: {e}")
        
        #Default moderate winter conditions if API fails
        return WeatherConditions(
            temperature_f=28.0,
            precipitation_type="snow",
            precipitation_rate=0.5,
            wind_speed_mph=15.0,
            visibility_miles=5.0,
            road_temperature_f=25.0,
            timestamp=datetime.now()
        )

class TrafficAPI:
    """Fetch real-time traffic data"""
    
    @staticmethod
    def get_traffic_data(road_segments: List[EnhancedRoad]) -> Dict[str, TrafficData]:
        """
        Get traffic data for road segments
        Note: Using simulated data based on road characteristics
        In production, integrate with HERE Traffic API, TomTom, or local traffic systems
        """
        traffic_data = {}
        
        for road in road_segments:
            # Simulate traffic based on road type and time
            hour = datetime.now().hour
            
            # Rush hour simulation
            is_rush_hour = (7 <= hour <= 9) or (16 <= hour <= 18)
            
            # Base congestion by road type
            if road.highway_type in ['motorway', 'trunk', 'primary']:
                base_congestion = 0.6 if is_rush_hour else 0.2
            elif road.highway_type in ['secondary', 'tertiary']:
                base_congestion = 0.4 if is_rush_hour else 0.15
            else:
                base_congestion = 0.2 if is_rush_hour else 0.05
            
            # Add randomness
            congestion = min(1.0, base_congestion + np.random.uniform(-0.1, 0.2))
            
            free_flow = road.maxspeed_mph if road.maxspeed_mph else 35.0
            current_speed = free_flow * (1 - congestion * 0.7)
            
            traffic_data[road.osm_id] = TrafficData(
                road_id=road.osm_id,
                current_speed_mph=current_speed,
                free_flow_speed_mph=free_flow,
                congestion_level=congestion,
                incident_nearby=np.random.random() < 0.05,  # 5% chance of incident
                timestamp=datetime.now()
            )
        
        return traffic_data

class PittsburghRoadNetwork:
    """Manage Pittsburgh road network from OpenStreetMap"""
    
    def __init__(self):
        self.graph: Optional[nx.MultiDiGraph] = None
        self.roads: Dict[str, EnhancedRoad] = {}
        self.node_coordinates: Dict[int, Tuple[float, float]] = {}
        self.priority_roads: Dict[int, List[EnhancedRoad]] = {1: [], 2: [], 3: [], 4: []}
    
    def should_include_road(self, road_name: str, highway_type: str) -> bool:
        """
        Determine if road is City of Pittsburgh responsibility
        Returns True if road should be included (city road), False if excluded (state/county)
        """
        if not CITY_ONLY_MODE:
            return True  # Include all roads if city-only mode disabled
        
        #Exclude major highways (PennDOT maintained)
        if highway_type in EXCLUDE_HIGHWAY_TYPES:
            return False
        
#Exclude known PennDOT roads by name
        if road_name:
            if isinstance(road_name, list):
                road_lower = ' '.join(road_name).lower()
            else:
                road_lower = road_name.lower()
            for penndot_road in PENNDOT_MAINTAINED_ROADS:
                if penndot_road.lower() in road_lower:
                    return False
        
        #Include all other roads (city responsibility)
        return True
        
    def load_from_osm(self, bbox: Tuple[float, float, float, float] = PITTSBURGH_BBOX):
        """Load Pittsburgh road network from OpenStreetMap"""
        print("Downloading Pittsburgh road network from OpenStreetMap...")
        
        #Download street network
        self.graph = ox.graph_from_place(
    "Pittsburgh, Pennsylvania, USA",
    network_type='drive',
    simplify=True
)
        
        print(f"Loaded {len(self.graph.nodes)} nodes and {len(self.graph.edges)} edges")
        
        #Store node coordinates
        for node, data in self.graph.nodes(data=True):
            self.node_coordinates[node] = (data['y'], data['x'])  # lat, lon
        
        #Convert to enhanced roads
        self._process_roads()
        
        #Classify priorities
        self._classify_road_priorities()
        
        print(f"Processed {len(self.roads)} road segments")
        print(f"  Priority 1 (Emergency): {len(self.priority_roads[1])}")
        print(f"  Priority 2 (High): {len(self.priority_roads[2])}")
        print(f"  Priority 3 (Medium): {len(self.priority_roads[3])}")
        print(f"  Priority 4 (Low): {len(self.priority_roads[4])}")
    
    def _process_roads(self):
        """Convert OSM graph to EnhancedRoad objects"""
        excluded_count = 0
        
        for u, v, key, data in self.graph.edges(keys=True, data=True):
            #Extract road attributes first for filtering
            highway_type = data.get('highway', 'unclassified')
            if isinstance(highway_type, list):
                highway_type = highway_type[0]
            
            name = data.get('name', f'Unnamed {highway_type}')
            
            #FILTER: Check if this road should be included (city responsibility)
            if not self.should_include_road(name, highway_type):
                excluded_count += 1
                continue  # Skip this road - not city responsibility
            
            #Create road segment
            osm_id = f"{u}_{v}_{key}"
            
            #Get geometry
            if 'geometry' in data:
                geometry = [(point[1], point[0]) for point in data['geometry'].coords]
            else:
                # Use straight line if no geometry
                geometry = [self.node_coordinates[u], self.node_coordinates[v]]
            
            #Extract remaining road attributes
            length = data.get('length', 100)  # meters
            
            #Parse lanes
            lanes = 2  # default
            if 'lanes' in data:
                try:
                    lanes = int(data['lanes']) if isinstance(data['lanes'], (int, float, str)) else 2
                except:
                    lanes = 2
            
            #Parse speed limit
            maxspeed = None
            if 'maxspeed' in data:
                try:
                    speed_str = str(data['maxspeed'])
                    if 'mph' in speed_str:
                        maxspeed = float(speed_str.replace('mph', '').strip())
                    else:
                        maxspeed = float(speed_str) * 0.621371  # Convert kph to mph
                except:
                    pass
            
            #Check for bridge/tunnel
            is_bridge = data.get('bridge', False) == 'yes' or data.get('bridge', False) is True
            is_tunnel = data.get('tunnel', False) == 'yes' or data.get('tunnel', False) is True
            
            #Calculate grade (elevation change)
            grade = 0.0
            if 'grade' in data:
                try:
                    grade = float(data['grade'])
                except:
                    pass
            
            road = EnhancedRoad(
                osm_id=osm_id,
                name=name,
                highway_type=highway_type,
                start_node=u,
                end_node=v,
                length_meters=length,
                geometry=geometry,
                lanes=lanes,
                maxspeed_mph=maxspeed,
                surface=data.get('surface'),
                priority=4,  # Will be updated in classification
                is_bridge=is_bridge,
                is_tunnel=is_tunnel,
                grade_percent=grade
            )
            
            self.roads[osm_id] = road
        
        #Report filtering results
        if CITY_ONLY_MODE:
            print(f"  Loaded {len(self.roads)} city roads (City of Pittsburgh DPW responsibility)")
            print(f"  Excluded {excluded_count} state/county roads (PennDOT/Allegheny County)")
        else:
            print(f"  Loaded {len(self.roads)} roads (all roads - city + state + county)")
    
    def _classify_road_priorities(self):
        """Classify roads by priority based on OpenStreetMap tags and proximity to critical infrastructure"""
        
        #Download points of interest
        print("Identifying critical infrastructure...")
        
        try:
            #Hospitals
            hospitals = ox.features_from_bbox(
                PITTSBURGH_BBOX[1], PITTSBURGH_BBOX[0], PITTSBURGH_BBOX[2], PITTSBURGH_BBOX[3],
                tags={'amenity': 'hospital'}
            )
            hospital_coords = [(geom.centroid.y, geom.centroid.x) for geom in hospitals.geometry if geom is not None]
            
            #Fire stations
            fire_stations = ox.features_from_bbox(
                PITTSBURGH_BBOX[1], PITTSBURGH_BBOX[0], PITTSBURGH_BBOX[2], PITTSBURGH_BBOX[3],
                tags={'amenity': 'fire_station'}
            )
            fire_coords = [(geom.centroid.y, geom.centroid.x) for geom in fire_stations.geometry if geom is not None]
            
            #Schools
            schools = ox.features_from_bbox(
                PITTSBURGH_BBOX[1], PITTSBURGH_BBOX[0], PITTSBURGH_BBOX[2], PITTSBURGH_BBOX[3],
                tags={'amenity': 'school'}
            )
            school_coords = [(geom.centroid.y, geom.centroid.x) for geom in schools.geometry if geom is not None]
            
            print(f"Found {len(hospital_coords)} hospitals, {len(fire_coords)} fire stations, {len(school_coords)} schools")
            
        except Exception as e:
            print(f"Warning: Could not download all POIs: {e}")
            hospital_coords = []
            fire_coords = []
            school_coords = []
        
        #Build KD trees for proximity searches
        hospital_tree = KDTree(hospital_coords) if hospital_coords else None
        fire_tree = KDTree(fire_coords) if fire_coords else None
        school_tree = KDTree(school_coords) if school_coords else None
        
        #Classify each road
        for road in self.roads.values():
            # Get road midpoint
            mid_idx = len(road.geometry) // 2
            road_coord = road.geometry[mid_idx]
            
            #Check proximity to critical infrastructure (within 0.005 degrees, ~500m)
            if hospital_tree:
                dist, _ = hospital_tree.query(road_coord)
                road.connects_hospital = dist < 0.005
            
            if fire_tree:
                dist, _ = fire_tree.query(road_coord)
                road.connects_fire_station = dist < 0.005
            
            if school_tree:
                dist, _ = school_tree.query(road_coord)
                road.connects_school = dist < 0.005
            
            # Priority 1: Emergency routes
            if road.highway_type in ['motorway', 'trunk'] or road.connects_hospital or road.connects_fire_station:
                road.priority = 1
                road.avg_daily_traffic = 30000
            
            #Priority 2: High traffic and important routes
            elif (road.highway_type in ['primary', 'primary_link'] or 
                  road.is_bridge or 
                  road.connects_school or
                  abs(road.grade_percent) > 5):
                road.priority = 2
                road.avg_daily_traffic = 15000
            
            #Priority 3: Secondary routes
            elif road.highway_type in ['secondary', 'tertiary', 'secondary_link', 'tertiary_link']:
                road.priority = 3
                road.avg_daily_traffic = 5000
            
            #Priority 4: Residential and low priority
            else:
                road.priority = 4
                road.avg_daily_traffic = 1000
            
            #Add to priority lists
            self.priority_roads[road.priority].append(road)
    
    def find_shortest_path(self, start_node: int, end_node: int) -> Tuple[List[int], float]:
        """Find shortest path between nodes using Dijkstra's algorithm"""
        try:
            path = nx.shortest_path(self.graph, start_node, end_node, weight='length')
            length = nx.shortest_path_length(self.graph, start_node, end_node, weight='length')
            return path, length / 1609.34  # Convert to miles
        except nx.NetworkXNoPath:
            return [], float('inf')
    
    def get_nearest_node(self, lat: float, lon: float) -> int:
        """Find nearest graph node to given coordinates"""
        return ox.distance.nearest_nodes(self.graph, lon, lat)

class SaltTruckOptimizer:
    """Main optimization engine"""
    
    def __init__(self, network: PittsburghRoadNetwork, depots: List[Depot], 
                 weather: WeatherConditions):
        self.network = network
        self.depots = depots
        self.weather = weather
        self.total_trucks = sum(d.available_trucks for d in depots)
        self.traffic_data: Dict[str, TrafficData] = {}
        
        #Create Truck objects for each depot
        self.trucks: List[Truck] = []
        truck_id = 1
        for depot in self.depots:
            for _ in range(depot.available_trucks):
                truck = Truck(
                    id=truck_id,
                    depot_id=depot.id,
                    tank_capacity_tons=12.0,
                    current_salt_tons=12.0,
                    driver_shift_hours=8.0
                )
                self.trucks.append(truck)
                truck_id += 1
        
        #Map depots to nearest nodes
        for depot in self.depots:
            depot.osm_node = self.network.get_nearest_node(depot.latitude, depot.longitude)
    
    def fetch_traffic_data(self):
        """Get current traffic conditions"""
        print("Fetching traffic data...")
        all_roads = list(self.network.roads.values())
        self.traffic_data = TrafficAPI.get_traffic_data(all_roads)
    
    def calculate_priority_score(self, road: EnhancedRoad) -> float:
        """
        Calculate comprehensive priority score.
        Lower score = higher priority (serviced first)
        """
        score = road.priority * 100
        
        #Critical infrastructure
        if road.connects_hospital:
            score -= 200
        if road.connects_fire_station:
            score -= 150
        if road.connects_school:
            score -= 80
        
        #Road characteristics
        if road.is_bridge:
            score -= 100  # Bridges ice first
        if abs(road.grade_percent) > 5:
            score -= 70  # Steep hills are dangerous
        if road.highway_type in ['motorway', 'trunk']:
            score -= 60
        
        #Traffic conditions
        if road.osm_id in self.traffic_data:
            traffic = self.traffic_data[road.osm_id]
            score += traffic.get_priority_boost()
        
        #Weather impact
        if self.weather.temperature_f < 20 and road.is_bridge:
            score -= 50  # Extra priority in extreme cold
        
        return score
    
    def create_truck_route(self, truck_id: int, depot: Depot, 
                          assigned_roads: List[EnhancedRoad]) -> Dict:
        """Create optimized route for single truck with refills, breaks, and shift management"""
        
        if not assigned_roads:
            return {
                "truck_id": truck_id,
                "depot": depot.name,
                "route": [],
                "total_time_minutes": 0,
                "total_distance_miles": 0,
                "roads_serviced": 0,
                "refills_needed": 0,
                "breaks_taken": 0,
                "salt_used_tons": 0,
                "shift_compliant": True
            }
        
        #Get the truck object
        truck = self.trucks[truck_id - 1]
        
        #Sort by priority score
        sorted_roads = sorted(assigned_roads, key=self.calculate_priority_score)
        
        route = []
        total_time = 0
        total_distance = 0
        current_node = depot.osm_node
        remaining_roads = set(r.osm_id for r in sorted_roads)
        refills_count = 0
        breaks_count = 0
        total_salt_used = 0
        
        #Greedy nearest-neighbor with priority weighting AND capacity constraints
        while remaining_roads:
            best_road = None
            best_score = float('inf')
            best_travel_dist = 0
            
            #Check if driver needs break
            if truck.needs_break() and len(route) > 0:
                break_time = 30.0
                truck.take_break(break_time)
                total_time += break_time
                breaks_count += 1
                
                route.append({
                    "segment_type": "break",
                    "description": "Mandatory driver break (30 min)",
                    "time_minutes": break_time,
                    "cumulative_time": total_time,
                    "salt_remaining_tons": round(truck.current_salt_tons, 2)
                })
            
            #Find best next road - SUPER OPTIMIZED: Use straight-line distance approximation
            #Get road objects for remaining roads
            candidate_roads = [self.network.roads[road_id] for road_id in remaining_roads]
            
            #Get current position
            current_pos = self.network.node_coordinates.get(current_node, (0, 0))
            
            #MUCH FASTER: Use straight-line distance instead of shortest_path
            #This is an approximation but 100x faster
            for road in candidate_roads:
                priority_score = self.calculate_priority_score(road)
                
                #Approximate travel distance using straight-line distance (Euclidean)
                road_start_pos = self.network.node_coordinates.get(road.start_node, road.geometry[0])
                travel_dist = ((road_start_pos[0] - current_pos[0])**2 + 
                              (road_start_pos[1] - current_pos[1])**2)**0.5
                
                #Convert lat/lon degrees to approximate miles (rough: 1 degree ≈ 69 miles)
                travel_dist_miles = travel_dist * 69
                
                #Combined score: prioritize high-priority roads, but consider distance
                combined_score = priority_score + (travel_dist_miles * 10)
                
                if combined_score < best_score:
                    best_road = road
                    best_score = combined_score
                    best_travel_dist = travel_dist_miles

            
            if best_road:
                #Check if refill needed BEFORE servicing this road
                salt_needed = truck.salt_needed_for_road(best_road)
                
                if not truck.can_service_road(best_road):
                    #Need to refill - find nearest depot
                    nearest_depot = depot  # Simplified: use home depot
                    refill_time = truck.refill_salt(depot.refill_time_minutes)
                    total_time += refill_time
                    refills_count += 1
                    
                    route.append({
                        "segment_type": "refill",
                        "description": f"Salt refill at {depot.name}",
                        "time_minutes": refill_time,
                        "cumulative_time": total_time,
                        "salt_loaded_tons": truck.tank_capacity_tons
                    })
                
                #Travel to road
                total_distance += best_travel_dist
                travel_time = best_travel_dist * 2  # Assume 30 mph travel
                total_time += travel_time
                truck.add_work_time(travel_time)
                
                #Service road
                service_time = best_road.salt_time_minutes(self.weather)
                total_time += service_time
                total_distance += best_road.length_meters / 1609.34
                truck.add_work_time(service_time)
                
                #Use salt
                truck.use_salt(salt_needed)
                total_salt_used += salt_needed / 2000  # Convert to tons
                
                route.append({
                    "segment_type": "road",
                    "road_id": best_road.osm_id,
                    "road_name": best_road.name,
                    "highway_type": best_road.highway_type,
                    "priority": best_road.priority,
                    "priority_score": self.calculate_priority_score(best_road),
                    "service_time_minutes": round(service_time, 1),
                    "length_miles": round(best_road.length_meters / 1609.34, 2),
                    "salt_used_lbs": round(salt_needed, 1),
                    "cumulative_time": round(total_time, 1),
                    "salt_remaining_tons": round(truck.current_salt_tons, 2),
                    "geometry": best_road.geometry
                })
                
                remaining_roads.remove(best_road.osm_id)
                current_node = best_road.end_node
                
                #Check if shift limit approaching
                if not truck.can_continue_shift(1.0):  # Less than 1 hour left
                    #End route - can't complete more roads in this shift
                    break
        
        #Calculate shift compliance
        shift_compliant = truck.current_shift_hours <= truck.driver_shift_hours
        shift_status = "✓ Within shift" if shift_compliant else f"⚠ {truck.current_shift_hours - truck.driver_shift_hours:.1f} hrs over"
        
        return {
            "truck_id": truck_id,
            "depot": depot.name,
            "depot_coords": (depot.latitude, depot.longitude),
            "route": route,
            "total_time_minutes": round(total_time, 1),
            "total_distance_miles": round(total_distance, 1),
            "roads_serviced": sum(1 for r in route if r.get("segment_type") == "road"),
            "refills_needed": refills_count,
            "breaks_taken": breaks_count,
            "salt_used_tons": round(total_salt_used, 2),
            "salt_remaining_tons": round(truck.current_salt_tons, 2),
            "shift_hours": round(truck.current_shift_hours, 2),
            "shift_compliant": shift_compliant,
            "shift_status": shift_status,
            "roads_not_completed": len(remaining_roads)
        }
    
    def optimize_fleet(self) -> Dict:
        """Main optimization algorithm - Optimized for large fleets (50-100 trucks)"""
        
        print("\n" + "="*70)
        print("OPTIMIZING SALT TRUCK FLEET DEPLOYMENT")
        print("="*70)
        
        #Fetch traffic data
        self.fetch_traffic_data()
        
        #Collect all roads prioritized
        all_roads = []
        for priority in [1, 2, 3, 4]:
            roads = self.network.priority_roads[priority]
            sorted_roads = sorted(roads, key=self.calculate_priority_score)
            all_roads.extend(sorted_roads)
        
        print(f"\nFleet Configuration:")
        print(f"  Total roads to service: {len(all_roads)}")
        print(f"  Available trucks: {self.total_trucks}")
        print(f"  Roads per truck (avg): {len(all_roads) / self.total_trucks:.1f}")
        print(f"  Depot facilities: {len(self.depots)}")
        
        #Assign roads to trucks using intelligent geographic clustering
        #Optimized for large fleets (50-100 trucks)
        truck_assignments = [[] for _ in range(self.total_trucks)]
        
        if self.total_trucks >= 20:
            #For large fleets, assign roads to nearest depot region first
            print(f"  Using depot-based geographic clustering for {self.total_trucks}-truck fleet")
            
            depot_regions = {depot.id: [] for depot in self.depots}
            depot_truck_ranges = {}
            
            #Calculate truck index range for each depot
            start_idx = 0
            for depot in self.depots:
                depot_truck_ranges[depot.id] = (start_idx, start_idx + depot.available_trucks)
                start_idx += depot.available_trucks
            
            # Assign each road to its nearest depot region
            for road in all_roads:
                mid_idx = len(road.geometry) // 2
                road_lat, road_lon = road.geometry[mid_idx]
                
                # Find nearest depot using Euclidean distance
                min_dist = float('inf')
                nearest_depot = None
                
                for depot in self.depots:
                    # Simple distance calculation (lat/lon differences)
                    dist = ((road_lat - depot.latitude)**2 + (road_lon - depot.longitude)**2)**0.5
                    if dist < min_dist:
                        min_dist = dist
                        nearest_depot = depot.id
                
                depot_regions[nearest_depot].append(road)
            
            # Distribute roads within each depot region to its trucks
            for depot_id, roads_in_region in depot_regions.items():
                truck_start, truck_end = depot_truck_ranges[depot_id]
                num_trucks_in_depot = truck_end - truck_start
                
                # Round-robin distribution within depot region
                for idx, road in enumerate(roads_in_region):
                    truck_idx = truck_start + (idx % num_trucks_in_depot)
                    truck_assignments[truck_idx].append(road)
                
                depot_name = next(d.name for d in self.depots if d.id == depot_id)
                print(f"    {depot_name}: {len(roads_in_region)} roads → {num_trucks_in_depot} trucks")
        else:
            # For smaller fleets, use simple round-robin
            print(f"  Using round-robin assignment for {self.total_trucks}-truck fleet")
            for idx, road in enumerate(all_roads):
                truck_assignments[idx % self.total_trucks].append(road)
        
        # Create routes
        print(f"\nGenerating optimized routes for {self.total_trucks} trucks...")
        routes = []
        truck_id = 1
        
        for depot in self.depots:
            for i in range(depot.available_trucks):
                if truck_assignments:
                    assigned = truck_assignments.pop(0)
                    route = self.create_truck_route(truck_id, depot, assigned)
                    routes.append(route)
                    
                    # Progress indicator for large fleets
                    if truck_id % 10 == 0:
                        print(f"  Completed {truck_id}/{self.total_trucks} routes...")
                    
                    truck_id += 1
        
        print(f"  ✓ All {self.total_trucks} routes generated!")
        
        #Calculate summary
        summary = self._create_summary(routes)
        
        return {
            "timestamp": datetime.now().isoformat(),
            "weather": {
                "temperature_f": self.weather.temperature_f,
                "precipitation": self.weather.precipitation_type,
                "severity_multiplier": self.weather.get_severity_multiplier()
            },
            "total_trucks_deployed": self.total_trucks,
            "routes": routes,
            "summary": summary
        }
    
    def _create_summary(self, routes: List[Dict]) -> Dict:
        """Create summary statistics"""
        if not routes:
            return {}
        
        total_roads = sum(r["roads_serviced"] for r in routes)
        max_time = max(r["total_time_minutes"] for r in routes)
        avg_time = np.mean([r["total_time_minutes"] for r in routes])
        total_distance = sum(r["total_distance_miles"] for r in routes)
        
        #New metrics: refills, breaks, salt usage, shift compliance
        total_refills = sum(r.get("refills_needed", 0) for r in routes)
        total_breaks = sum(r.get("breaks_taken", 0) for r in routes)
        total_salt_used = sum(r.get("salt_used_tons", 0) for r in routes)
        
        shift_compliant_count = sum(1 for r in routes if r.get("shift_compliant", True))
        shift_compliance_pct = (shift_compliant_count / len(routes)) * 100 if routes else 0
        
        roads_not_completed = sum(r.get("roads_not_completed", 0) for r in routes)
        
        #Priority breakdown
        priority_counts = {1: 0, 2: 0, 3: 0, 4: 0}
        for route in routes:
            for segment in route["route"]:
                if segment.get("segment_type") == "road":
                    priority_counts[segment["priority"]] += 1
        
        return {
            "total_roads_serviced": total_roads,
            "roads_not_completed": roads_not_completed,
            "estimated_completion_time_hours": round(max_time / 60, 2),
            "average_truck_time_hours": round(avg_time / 60, 2),
            "total_fleet_distance_miles": round(total_distance, 1),
            "priority_breakdown": priority_counts,
            "efficiency_score": "optimal" if np.std([r["total_time_minutes"] for r in routes]) < 30 else "needs_balancing",
            "total_salt_refills": total_refills,
            "total_driver_breaks": total_breaks,
            "total_salt_used_tons": round(total_salt_used, 1),
            "avg_salt_per_truck_tons": round(total_salt_used / len(routes), 2),
            "shift_compliance_percent": round(shift_compliance_pct, 1),
            "trucks_over_shift": len(routes) - shift_compliant_count
        }

class RouteVisualizer:
    """Create map visualizations of routes"""
    
    @staticmethod
    def create_route_map(results: Dict, output_file: str = "C:/Users/JeJones/Desktop/pittsburgh_salt_routes.html"):
        """Create interactive Folium map with all routes"""
        
        #Create base map centered on Pittsburgh
        m = folium.Map(
            location=PITTSBURGH_COORDS,
            zoom_start=12,
            tiles='OpenStreetMap'
        )
        
        #Color scheme for priorities
        priority_colors = {
            1: '#FF0000',  # Red - Emergency
            2: '#FF8C00',  # Orange - High
            3: '#FFD700',  # Yellow - Medium
            4: '#90EE90'   # Light Green - Low
        }
        
        #Add routes
        for route in results['routes']:
            depot_coords = route['depot_coords']
            
            # Add depot marker
            folium.Marker(
                location=depot_coords,
                popup=f"Depot: {route['depot']}<br>Truck #{route['truck_id']}<br>Roads: {route['roads_serviced']}",
                icon=folium.Icon(color='blue', icon='home')
            ).add_to(m)
            
            #Add route segments
            for segment in route['route']:
                seg_type = segment.get('segment_type', 'road')
                
                if seg_type == 'road':
                    color = priority_colors.get(segment['priority'], '#808080')
                    
                    folium.PolyLine(
                        locations=segment['geometry'],
                        color=color,
                        weight=3,
                        opacity=0.7,
                        popup=f"{segment['road_name']}<br>Priority: {segment['priority']}<br>Time: {segment['service_time_minutes']} min<br>Salt: {segment.get('salt_used_lbs', 0):.0f} lbs"
                    ).add_to(m)
                
                elif seg_type == 'refill':
                    # Add refill marker
                    folium.CircleMarker(
                        location=depot_coords,
                        radius=8,
                        popup=f"🔄 Refill Stop<br>{segment['description']}<br>Time: {segment['time_minutes']} min",
                        color='orange',
                        fill=True,
                        fillColor='orange',
                        fillOpacity=0.8
                    ).add_to(m)
        
        #Add legend
        legend_html = '''
        <div style="position: fixed; 
                    bottom: 50px; right: 50px; width: 200px; height: 180px; 
                    background-color: white; border:2px solid grey; z-index:9999; 
                    font-size:14px; padding: 10px">
        <p><strong>Route Priorities</strong></p>
        <p><span style="color: #FF0000;">━━━</span> Priority 1: Emergency</p>
        <p><span style="color: #FF8C00;">━━━</span> Priority 2: High</p>
        <p><span style="color: #FFD700;">━━━</span> Priority 3: Medium</p>
        <p><span style="color: #90EE90;">━━━</span> Priority 4: Low</p>
        <p><span style="color: #0000FF;">📍</span> Depot Locations</p>
        </div>
        '''
        m.get_root().html.add_child(folium.Element(legend_html))
        
        # Save map
        m.save(output_file)
        print(f"\nInteractive map saved to: {output_file}")
        
        return output_file

def main():
    """Main execution"""
    
    print("="*70)
    print("PITTSBURGH WINTER STORM SALT TRUCK OPTIMIZATION SYSTEM")
    print("Real-time Integration: OpenStreetMap + Weather + Traffic")
    if CITY_ONLY_MODE:
        print("Mode: CITY OF PITTSBURGH ROADS ONLY (DPW Responsibility)")
        print("    Excludes PennDOT state highways and county roads")
    else:
        print("Mode: ALL ROADS (City + State + County)")
    print("="*70)
    print()
    
    #Step 1: Get weather conditions
    print("Step 1: Fetching current weather conditions...")
    weather = WeatherAPI.get_current_conditions(*PITTSBURGH_COORDS)
    print(f"  Temperature: {weather.temperature_f}°F")
    print(f"  Precipitation: {weather.precipitation_type}")
    print(f"  Severity Multiplier: {weather.get_severity_multiplier():.2f}x")
    print()
    
    #Step 2: Load road network
    print("Step 2: Loading Pittsburgh road network from OpenStreetMap...")
    network = PittsburghRoadNetwork()
    network.load_from_osm()
    print()
    
    #Step 3: Define depots (major Pittsburgh DPW locations)
    #Fleet size adjusted based on city-only vs all roads
    #City-only: ~900 roads → 65-70 trucks optimal
    #All roads: ~1,200+ roads → 90 trucks
    
    if CITY_ONLY_MODE:
        # Optimized for ~900 city roads
        fleet_multiplier = 0.78  # 78% of full fleet (90 → 70 trucks)
    else:
        # Full fleet for all roads
        fleet_multiplier = 1.0  # 100% (90 trucks)
    
    depots = [
        Depot(
            id="depot_highland",
            name="Highland Park DPW Facility (Main)",
            latitude=40.4805,
            longitude=-79.9140,
            available_trucks=int(25 * fleet_multiplier),  # 25 or 20
            salt_capacity_tons=250
        ),
        Depot(
            id="depot_south",
            name="South Side Works Facility",
            latitude=40.4278,
            longitude=-79.9789,
            available_trucks=int(20 * fleet_multiplier),  # 20 or 16
            salt_capacity_tons=200
        ),
        Depot(
            id="depot_north",
            name="North Side Facility",
            latitude=40.4564,
            longitude=-80.0152,
            available_trucks=int(18 * fleet_multiplier),  # 18 or 14
            salt_capacity_tons=180
        ),
        Depot(
            id="depot_west",
            name="West End Facility",
            latitude=40.4423,
            longitude=-80.0423,
            available_trucks=int(15 * fleet_multiplier),  # 15 or 12
            salt_capacity_tons=150
        ),
        Depot(
            id="depot_east",
            name="East Liberty Facility",
            latitude=40.4615,
            longitude=-79.9214,
            available_trucks=int(12 * fleet_multiplier),  # 12 or 9
            salt_capacity_tons=120
        )
    ]
    
    total_trucks = sum(d.available_trucks for d in depots)
    print(f"  Total Fleet Size: {total_trucks} trucks")
    if CITY_ONLY_MODE:
        print(f"  Optimized for City of Pittsburgh roads (~900 centerline miles)")
    else:
        print(f"  Configured for all roads (city + state + county)")
    
    print("Step 3: Depot Configuration")
    for depot in depots:
        print(f"  {depot.name}: {depot.available_trucks} trucks")
    print()
    
    #Step 4: Run optimization
    print("Step 4: Running optimization algorithm...")
    optimizer = SaltTruckOptimizer(network, depots, weather)
    results = optimizer.optimize_fleet()
    print()
    
    #Step 5: Display results
    print("="*70)
    print("OPTIMIZATION RESULTS")
    print("="*70)
    print()
    print(f"Deployment Time: {results['timestamp']}")
    print(f"Weather Conditions: {results['weather']['temperature_f']}°F, {results['weather']['precipitation']}")
    print(f"Weather Severity: {results['weather']['severity_multiplier']:.2f}x normal")
    print()
    print(f"Fleet Deployment:")
    print(f"  Total Trucks: {results['total_trucks_deployed']}")
    print(f"  Total Roads: {results['summary']['total_roads_serviced']}")
    print(f"  Roads Not Completed: {results['summary'].get('roads_not_completed', 0)}")
    print(f"  Estimated Completion: {results['summary']['estimated_completion_time_hours']:.1f} hours")
    print(f"  Total Fleet Distance: {results['summary']['total_fleet_distance_miles']:.1f} miles")
    print(f"  Average Truck Time: {results['summary']['average_truck_time_hours']:.1f} hours")
    print()
    print(f"Operational Metrics:")
    print(f"  Total Salt Used: {results['summary'].get('total_salt_used_tons', 0):.1f} tons")
    print(f"  Avg Salt per Truck: {results['summary'].get('avg_salt_per_truck_tons', 0):.2f} tons")
    print(f"  Salt Refills Required: {results['summary'].get('total_salt_refills', 0)}")
    print(f"  Driver Breaks Taken: {results['summary'].get('total_driver_breaks', 0)}")
    print(f"  Shift Compliance: {results['summary'].get('shift_compliance_percent', 100):.1f}%")
    if results['summary'].get('trucks_over_shift', 0) > 0:
        print(f"  ⚠ Trucks Over Shift: {results['summary'].get('trucks_over_shift', 0)}")
    print()
    print(f"Priority Distribution:")
    for priority, count in results['summary']['priority_breakdown'].items():
        priority_names = {1: "Emergency", 2: "High", 3: "Medium", 4: "Low"}
        print(f"  Priority {priority} ({priority_names[priority]}): {count} roads")
    print()
    
    #Show sample routes
    print("="*70)
    print(f"SAMPLE TRUCK ROUTES (First 5 of {results['total_trucks_deployed']} Trucks)")
    print("="*70)
    
    for route in results['routes'][:5]:
        print(f"\nTruck #{route['truck_id']} - {route['depot']}")
        print(f"  Total Time: {route['total_time_minutes']:.0f} min ({route['total_time_minutes']/60:.1f} hrs)")
        print(f"  Total Distance: {route['total_distance_miles']:.1f} miles")
        print(f"  Roads: {route['roads_serviced']}")
        print(f"  Salt Used: {route.get('salt_used_tons', 0):.2f} tons | Remaining: {route.get('salt_remaining_tons', 0):.2f} tons")
        print(f"  Refills: {route.get('refills_needed', 0)} | Breaks: {route.get('breaks_taken', 0)}")
        print(f"  Shift: {route.get('shift_hours', 0):.1f} hrs | Status: {route.get('shift_status', 'Unknown')}")
        print(f"  Route Details (first 6 segments):")
        
        segment_count = 0
        for idx, segment in enumerate(route['route'], 1):
            if segment_count >= 6:
                break
                
            seg_type = segment.get("segment_type", "road")
            
            if seg_type == "road":
                priority_names = {1: "EMERGENCY", 2: "HIGH", 3: "MEDIUM", 4: "LOW"}
                print(f"    {idx}. {segment['road_name'][:35]:<35} | P{segment['priority']} {priority_names[segment['priority']]:<9} | {segment['length_miles']:.2f}mi | {segment['service_time_minutes']:.0f}min | {segment.get('salt_used_lbs', 0):.0f}lbs")
                segment_count += 1
            elif seg_type == "refill":
                print(f"    {idx}. 🔄 REFILL STOP - {segment['description']:<20} | {segment['time_minutes']:.0f}min | Loaded: {segment.get('salt_loaded_tons', 0):.1f} tons")
                segment_count += 1
            elif seg_type == "break":
                print(f"    {idx}. ☕ DRIVER BREAK - {segment['description']:<20} | {segment['time_minutes']:.0f}min")
                segment_count += 1
        
        total_segments = len(route['route'])
        if total_segments > 6:
            print(f"    ... and {total_segments - 6} more segments")
    
    if results['total_trucks_deployed'] > 5:
        print(f"\n... and {results['total_trucks_deployed'] - 5} more trucks")
        print(f"\nComplete route data for all {results['total_trucks_deployed']} trucks saved to:")
        print(f"  • optimization_results.json")
        print(f"  • pittsburgh_salt_routes.html (interactive map)")
    
    
    print("\n" + "="*70)
    print("CREATING INTERACTIVE MAP VISUALIZATION")
    print("="*70)
    
    map_file = RouteVisualizer.create_route_map(results)
    
    #Calculate multi-pass strategy if heavy storm
    print("\n" + "="*70)
    print("STORM MANAGEMENT ANALYSIS")
    print("="*70)
    
    precip_rate = weather.precipitation_rate
    temp = weather.temperature_f
    
    #Determine if multiple passes needed
    if precip_rate >= 1.5:
        passes_needed = 3
        interval_hours = 2.0
        storm_severity = "HEAVY"
    elif precip_rate >= 0.75:
        passes_needed = 2
        interval_hours = 3.0
        storm_severity = "MODERATE"
    else:
        passes_needed = 1
        interval_hours = 0
        storm_severity = "LIGHT"
    
    print(f"Current Conditions: {storm_severity} {weather.precipitation_type}, {precip_rate:.1f}\"/hr, {temp}°F")
    print(f"Passes Required: {passes_needed}")
    
    if passes_needed > 1:
        print(f"\nMulti-Pass Strategy:")
        for pass_num in range(1, passes_needed + 1):
            start_hour = (pass_num - 1) * interval_hours
            if pass_num <= 2:
                print(f"  Pass {pass_num} (Hour {start_hour:.0f}): Priority 1 & 2 roads only - All {results['total_trucks_deployed']} trucks")
            else:
                print(f"  Pass {pass_num} (Hour {start_hour:.0f}): Priority 3 & 4 roads - All {results['total_trucks_deployed']} trucks")
        
        total_passes_time = results['summary']['estimated_completion_time_hours'] * passes_needed
        print(f"\nEstimated Total Storm Duration: {total_passes_time:.1f} hours")
        print(f"Total Salt Usage (all passes): {results['summary']['total_salt_used_tons'] * passes_needed:.1f} tons")
        print(f"⚠ Heavy storm detected - Continuous re-treatment required")
    else:
        print(f"\nSingle Pass Sufficient")
        print(f"Estimated completion: {results['summary']['estimated_completion_time_hours']:.1f} hours")
    
    #Salt budget warning
    total_salt_all_passes = results['summary']['total_salt_used_tons'] * passes_needed
    if total_salt_all_passes > 500:
        print(f"\n⚠ WARNING: High salt usage projected ({total_salt_all_passes:.0f} tons)")
        print(f"   Verify adequate supply at all depots")
    
    # Shift management warning  
    if results['summary'].get('trucks_over_shift', 0) > 0:
        print(f"\n⚠ WARNING: {results['summary']['trucks_over_shift']} trucks exceed 8-hour shift")
        print(f"   Recommend: Split routes or arrange relief drivers")
    
    print()
    
    #Save JSON results with custom encoder for numpy types
    def convert_numpy(obj):
        """Convert numpy types to native Python types for JSON serialization"""
        if isinstance(obj, (np.integer, np.floating)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, np.bool_):
            return bool(obj)
        elif isinstance(obj, dict):
            return {k: convert_numpy(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [convert_numpy(item) for item in obj]
        return obj
    
    json_file = "C:/Users/JeJones/Desktop/optimization_results.json"
    with open(json_file, 'w') as f:
        json.dump(convert_numpy(results), f, indent=2)
    
    print(f"Results saved to: {json_file}")
    
    print("\n" + "="*70)
    print("OPTIMIZATION COMPLETE")
    print("="*70)
    print("""
ENHANCED SYSTEM FEATURES:
✓ Real-time weather integration (National Weather Service)
✓ OpenStreetMap road network data
✓ Traffic condition simulation
✓ Priority-based routing (Emergency → High → Medium → Low)
✓ Geographic optimization
✓ Interactive route visualization
✓ Hospital/fire station/school proximity detection
✓ Bridge and hill identification
✓ Weather-adjusted service times

NEW OPERATIONAL ENHANCEMENTS:
✓ Salt tank capacity tracking (12 tons per truck)
✓ Automatic refill stop insertion
✓ Driver shift management (8-hour shifts)
✓ Mandatory break scheduling (every 4 hours)
✓ Shift compliance validation
✓ Real-time salt usage tracking
✓ Multi-pass strategy for heavy storms
✓ Salt budget projections
✓ Jurisdiction filtering (city roads only)

JURISDICTION MODE:
• {}
• {}

OPERATIONAL SUMMARY:
• Total Salt Used: {:.1f} tons
• Refill Stops Required: {}
• Driver Breaks: {}
• Shift Compliance: {:.0f}%
• Multiple Passes Needed: {}

NEXT STEPS:
1. Review interactive map and detailed routes
2. Verify salt supply at all depots
3. Brief drivers on refill locations and break times
4. Monitor weather for changes requiring re-optimization
5. Deploy fleet when ready
6. Track actual vs. planned performance
    """.format(
        "City of Pittsburgh roads ONLY (DPW responsibility)" if CITY_ONLY_MODE else "All roads (city + state + county)",
        "Excludes PennDOT highways and county roads" if CITY_ONLY_MODE else "Includes interstates, US routes, and county roads",
        results['summary'].get('total_salt_used_tons', 0),
        results['summary'].get('total_salt_refills', 0),
        results['summary'].get('total_driver_breaks', 0),
        results['summary'].get('shift_compliance_percent', 100),
        passes_needed
    ))

if __name__ == "__main__":
    main()
"""
Enhanced Map Generator for Pittsburgh Salt Truck Routes
Creates multiple map visualizations for different audiences
Run this AFTER PGH_Salt_Trucks_FINAL.py
"""

import json
import folium
from folium import plugins

PITTSBURGH_COORDS = (40.4406, -79.9959)

def generate_all_maps(json_file="C:/Users/JeJones/Desktop/optimization_results.json"):
    """Generate all map types"""
    
    print("="*70)
    print("GENERATING MULTIPLE MAP VISUALIZATIONS")
    print("="*70)
    
    print("\nLoading optimization results...")
    with open(json_file, 'r') as f:
        results = json.load(f)
    
    maps_created = []
    
    # Map 1: Full Fleet Map (detailed)
    print("\n1. Creating Full Fleet Map (all trucks, all routes)...")
    map1 = create_full_fleet_map(results)
    maps_created.append(map1)
    
    # Map 2: Priority Roads Only
    print("2. Creating Priority Roads Map (high priority only)...")
    map2 = create_priority_map(results)
    maps_created.append(map2)
    
    # Map 3: Public Simplified Map
    print("3. Creating Public Map (simplified for residents)...")
    map3 = create_public_map(results)
    maps_created.append(map3)
    
    # Map 4: Depot-Specific Maps
    print("4. Creating Depot-Specific Maps...")
    depot_maps = create_depot_maps(results)
    maps_created.extend(depot_maps)
    
    # Map 5: Heat Map (coverage intensity)
    print("5. Creating Coverage Heat Map...")
    map5 = create_heat_map(results)
    maps_created.append(map5)
    
    print("\n" + "="*70)
    print("MAP GENERATION COMPLETE")
    print("="*70)
    print(f"\nCreated {len(maps_created)} map files:")
    for map_file in maps_created:
        print(f"  • {map_file}")
    
    return maps_created

def create_full_fleet_map(results):
    """Detailed map with all trucks and routes"""
    output_file = "C:/Users/JeJones/Desktop/map_full_fleet.html"
    
    m = folium.Map(
        location=PITTSBURGH_COORDS,
        zoom_start=12,
        tiles='OpenStreetMap'
    )
    
    # Color palette for trucks
    colors = ['red', 'blue', 'green', 'purple', 'orange', 'darkred', 
              'lightred', 'beige', 'darkblue', 'darkgreen', 'cadetblue',
              'darkpurple', 'pink', 'lightblue', 'lightgreen', 'gray', 'black']
    
    routes = results.get('routes', [])
    
    # Add routes
    for idx, route in enumerate(routes):
        color = colors[idx % len(colors)]
        truck_id = route.get('truck_id')
        depot = route.get('depot', 'Unknown')
        
        # Create feature group for this truck
        truck_group = folium.FeatureGroup(name=f"Truck #{truck_id} - {depot}")
        
        # Add depot marker
        depot_coords = route.get('depot_coords')
        if depot_coords:
            folium.Marker(
                depot_coords,
                popup=f"<b>{depot}</b><br>Truck #{truck_id} Home Base",
                icon=folium.Icon(color='red', icon='home', prefix='fa'),
                tooltip=f"Truck #{truck_id} - {depot}"
            ).add_to(truck_group)
        
        # Add route segments
        for segment in route.get('route', []):
            if segment.get('segment_type') == 'road' and segment.get('geometry'):
                geometry = segment['geometry']
                priority = segment.get('priority', 4)
                road_name = segment.get('road_name', 'Unnamed')
                
                # Line weight based on priority
                weight = {1: 5, 2: 4, 3: 3, 4: 2}.get(priority, 2)
                
                folium.PolyLine(
                    geometry,
                    color=color,
                    weight=weight,
                    opacity=0.7,
                    popup=f"<b>{road_name}</b><br>Truck #{truck_id}<br>Priority {priority}",
                    tooltip=road_name
                ).add_to(truck_group)
        
        truck_group.add_to(m)
    
    # Add layer control
    folium.LayerControl(collapsed=False).add_to(m)
    
    # Add title
    title_html = '''
    <div style="position: fixed; top: 10px; left: 50px; z-index:9999; 
                background-color: white; padding: 10px; border: 2px solid gray; border-radius: 5px;">
        <h3 style="margin:0">Pittsburgh Salt Truck Fleet - Full Detail</h3>
        <p style="margin:0; font-size:12px">All {} trucks - Toggle layers to view individual routes</p>
    </div>
    '''.format(len(routes))
    m.get_root().html.add_child(folium.Element(title_html))
    
    m.save(output_file)
    return output_file

def create_priority_map(results):
    """Map showing only high-priority roads"""
    output_file = "C:/Users/JeJones/Desktop/map_priority_roads.html"
    
    m = folium.Map(
        location=PITTSBURGH_COORDS,
        zoom_start=12,
        tiles='OpenStreetMap'
    )
    
    priority_colors = {
        1: 'red',      # Emergency
        2: 'orange',   # High
        3: 'yellow',   # Medium
        4: 'gray'      # Low (not shown)
    }
    
    routes = results.get('routes', [])
    
    # Group by priority
    for priority in [1, 2, 3]:
        priority_name = {1: 'P1: Emergency', 2: 'P2: High Priority', 3: 'P3: Medium Priority'}[priority]
        priority_group = folium.FeatureGroup(name=priority_name)
        
        for route in routes:
            for segment in route.get('route', []):
                if (segment.get('segment_type') == 'road' and 
                    segment.get('priority') == priority and 
                    segment.get('geometry')):
                    
                    road_name = segment.get('road_name', 'Unnamed')
                    geometry = segment['geometry']
                    
                    folium.PolyLine(
                        geometry,
                        color=priority_colors[priority],
                        weight=4 if priority <= 2 else 2,
                        opacity=0.8,
                        popup=f"<b>{road_name}</b><br>Priority {priority}",
                        tooltip=road_name
                    ).add_to(priority_group)
        
        priority_group.add_to(m)
    
    folium.LayerControl().add_to(m)
    
    # Add legend
    legend_html = '''
    <div style="position: fixed; bottom: 50px; left: 50px; z-index:9999; 
                background-color: white; padding: 15px; border: 2px solid gray; border-radius: 5px;">
        <h4 style="margin-top:0">Priority Legend</h4>
        <p><span style="color:red">━━━</span> Priority 1: Emergency Routes</p>
        <p><span style="color:orange">━━━</span> Priority 2: High Priority</p>
        <p><span style="color:yellow">━━━</span> Priority 3: Medium Priority</p>
    </div>
    '''
    m.get_root().html.add_child(folium.Element(legend_html))
    
    m.save(output_file)
    return output_file

def create_public_map(results):
    """Simplified map for public/residents"""
    output_file = "C:/Users/JeJones/Desktop/map_public_simple.html"
    
    m = folium.Map(
        location=PITTSBURGH_COORDS,
        zoom_start=11,
        tiles='CartoDB positron'  # Cleaner tiles for public
    )
    
    routes = results.get('routes', [])
    summary = results.get('summary', {})
    
    # Just show covered areas (all roads as one color)
    all_roads = folium.FeatureGroup(name="Covered Roads")
    
    for route in routes:
        for segment in route.get('route', []):
            if segment.get('segment_type') == 'road' and segment.get('geometry'):
                geometry = segment['geometry']
                
                folium.PolyLine(
                    geometry,
                    color='blue',
                    weight=2,
                    opacity=0.4
                ).add_to(all_roads)
    
    all_roads.add_to(m)
    
    # Add depots only
    depot_locations = {
        'Highland Park DPW Facility': (40.4718, -79.9088),
        'South Side Works Facility': (40.4288, -79.9628),
        'North Side Facility': (40.4550, -80.0100),
        'West End Facility': (40.4391, -80.0580),
        'East Liberty Facility': (40.4615, -79.9245)
    }
    
    for depot_name, coords in depot_locations.items():
        folium.Marker(
            coords,
            popup=f"<b>{depot_name}</b>",
            icon=folium.Icon(color='red', icon='home', prefix='fa'),
            tooltip=depot_name
        ).add_to(m)
    
    # Add info box
    info_html = '''
    <div style="position: fixed; top: 10px; left: 50px; z-index:9999; 
                background-color: white; padding: 15px; border: 2px solid #3186cc; border-radius: 5px;
                max-width: 350px;">
        <h3 style="margin-top:0; color:#3186cc">Pittsburgh Winter Salt Truck Coverage</h3>
        <p><strong>Estimated Completion:</strong> {} hours</p>
        <p><strong>Total Roads:</strong> {}</p>
        <p><strong>Fleet Size:</strong> {} trucks from 5 depots</p>
        <p style="font-size:12px; color:gray; margin-top:10px">
            Blue lines show roads being treated. Trucks start from red depot markers.
        </p>
    </div>
    '''.format(
        summary.get('estimated_completion_time_hours', 'N/A'),
        summary.get('total_roads_serviced', 'N/A'),
        results.get('total_trucks_deployed', 'N/A')
    )
    m.get_root().html.add_child(folium.Element(info_html))
    
    m.save(output_file)
    return output_file

def create_depot_maps(results):
    """Create separate map for each depot"""
    depot_files = []
    routes = results.get('routes', [])
    
    # Group routes by depot
    depot_routes = {}
    for route in routes:
        depot = route.get('depot', 'Unknown')
        if depot not in depot_routes:
            depot_routes[depot] = []
        depot_routes[depot].append(route)
    
    # Create map for each depot
    for depot_name, depot_route_list in depot_routes.items():
        safe_name = depot_name.replace(' ', '_').replace('(', '').replace(')', '')
        output_file = f"C:/Users/JeJones/Desktop/map_depot_{safe_name}.html"
        
        m = folium.Map(
            location=PITTSBURGH_COORDS,
            zoom_start=12,
            tiles='OpenStreetMap'
        )
        
        # Add depot marker
        if depot_route_list:
            depot_coords = depot_route_list[0].get('depot_coords')
            if depot_coords:
                folium.Marker(
                    depot_coords,
                    popup=f"<b>{depot_name}</b><br>{len(depot_route_list)} trucks",
                    icon=folium.Icon(color='red', icon='home', prefix='fa', icon_size=(30, 30)),
                    tooltip=depot_name
                ).add_to(m)
        
        # Add routes
        colors = ['blue', 'green', 'purple', 'orange', 'darkred', 'cadetblue']
        for idx, route in enumerate(depot_route_list):
            color = colors[idx % len(colors)]
            truck_id = route.get('truck_id')
            
            for segment in route.get('route', []):
                if segment.get('segment_type') == 'road' and segment.get('geometry'):
                    geometry = segment['geometry']
                    road_name = segment.get('road_name', 'Unnamed')
                    
                    folium.PolyLine(
                        geometry,
                        color=color,
                        weight=3,
                        opacity=0.7,
                        popup=f"<b>{road_name}</b><br>Truck #{truck_id}",
                        tooltip=f"Truck #{truck_id}: {road_name}"
                    ).add_to(m)
        
        # Add title
        title_html = f'''
        <div style="position: fixed; top: 10px; left: 50px; z-index:9999; 
                    background-color: white; padding: 10px; border: 2px solid gray; border-radius: 5px;">
            <h3 style="margin:0">{depot_name}</h3>
            <p style="margin:0; font-size:12px">{len(depot_route_list)} trucks deployed</p>
        </div>
        '''
        m.get_root().html.add_child(folium.Element(title_html))
        
        m.save(output_file)
        depot_files.append(output_file)
    
    return depot_files

def create_heat_map(results):
    """Heat map showing coverage density"""
    output_file = "C:/Users/JeJones/Desktop/map_coverage_heatmap.html"
    
    m = folium.Map(
        location=PITTSBURGH_COORDS,
        zoom_start=12,
        tiles='CartoDB dark_matter'
    )
    
    # Collect all road points
    heat_data = []
    routes = results.get('routes', [])
    
    for route in routes:
        for segment in route.get('route', []):
            if segment.get('segment_type') == 'road' and segment.get('geometry'):
                geometry = segment['geometry']
                # Sample points along the road
                for i in range(0, len(geometry), max(1, len(geometry)//5)):  # Sample 5 points per road
                    heat_data.append(geometry[i])
    
    # Add heatmap
    if heat_data:
        plugins.HeatMap(heat_data, radius=15, blur=20, max_zoom=13).add_to(m)
    
    # Add title
    title_html = '''
    <div style="position: fixed; top: 10px; left: 50px; z-index:9999; 
                background-color: rgba(0,0,0,0.7); padding: 10px; border-radius: 5px; color: white;">
        <h3 style="margin:0">Coverage Density Heat Map</h3>
        <p style="margin:0; font-size:12px">Brighter areas show higher coverage density</p>
    </div>
    '''
    m.get_root().html.add_child(folium.Element(title_html))
    
    m.save(output_file)
    return output_file

if __name__ == "__main__":
    try:
        maps = generate_all_maps()
        print("\n✓ All maps generated successfully!")
        print("\nOpen these files in your web browser to view:")
        for map_file in maps:
            print(f"  {map_file}")
    except FileNotFoundError:
        print("\nERROR: optimization_results.json not found!")
        print("Please run PGH_Salt_Trucks_FINAL.py first.")
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
"""
Excel Report Generator for Pittsburgh Salt Truck Routes
Run this AFTER PGH_Salt_Trucks_FINAL.py to generate Excel reports
"""

import json
import sys

# Check if openpyxl is installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed!")
    print("Please run: pip install openpyxl")
    sys.exit(1)

def create_excel_report(json_file="C:/Users/JeJones/Desktop/optimization_results.json",
                       output_file="C:/Users/JeJones/Desktop/salt_truck_report.xlsx"):
    """Create comprehensive Excel workbook from JSON results"""
    
    print("Loading optimization results...")
    with open(json_file, 'r') as f:
        results = json.load(f)
    
    print("Creating Excel workbook...")
    wb = Workbook()
    
    # Sheet 1: Executive Summary
    print("  Creating Executive Summary...")
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    create_summary_sheet(ws_summary, results)
    
    # Sheet 2: All Truck Routes
    print("  Creating All Truck Routes...")
    ws_routes = wb.create_sheet("All Truck Routes")
    create_routes_sheet(ws_routes, results)
    
    # Sheet 3: Priority Analysis
    print("  Creating Priority Analysis...")
    ws_priority = wb.create_sheet("Priority Analysis")
    create_priority_sheet(ws_priority, results)
    
    # Sheet 4: Cost Analysis
    print("  Creating Cost Analysis...")
    ws_cost = wb.create_sheet("Cost Analysis")
    create_cost_sheet(ws_cost, results)
    
    # Sheet 5: Depot Performance
    print("  Creating Depot Performance...")
    ws_depot = wb.create_sheet("Depot Performance")
    create_depot_sheet(ws_depot, results)
    
    # Sheet 6: Detailed Route Segments
    print("  Creating Detailed Routes...")
    ws_detailed = wb.create_sheet("Detailed Routes")
    create_detailed_routes_sheet(ws_detailed, results)
    
    print(f"Saving Excel file...")
    wb.save(output_file)
    print(f"✓ Excel report saved to: {output_file}")
    return output_file

def create_summary_sheet(ws, results):
    """Executive summary with key metrics"""
    # Title
    ws['A1'] = 'PITTSBURGH SALT TRUCK FLEET - EXECUTIVE SUMMARY'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Timestamp
    ws['A2'] = f"Generated: {results.get('timestamp', 'N/A')}"
    ws['A2'].font = Font(italic=True)
    
    # Weather Info
    ws['A4'] = 'WEATHER CONDITIONS'
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A4'].font = Font(bold=True, color='FFFFFF', size=12)
    
    weather = results.get('weather', {})
    ws['A5'] = 'Temperature:'
    ws['B5'] = f"{weather.get('temperature_f', 'N/A')}°F"
    ws['A6'] = 'Precipitation:'
    ws['B6'] = weather.get('precipitation', 'N/A')
    ws['A7'] = 'Severity Multiplier:'
    ws['B7'] = f"{weather.get('severity_multiplier', 1.0)}x"
    
    # Fleet Summary
    ws['A9'] = 'FLEET DEPLOYMENT'
    ws['A9'].font = Font(bold=True, size=12)
    ws['A9'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A9'].font = Font(bold=True, color='FFFFFF', size=12)
    
    summary = results.get('summary', {})
    ws['A10'] = 'Total Trucks:'
    ws['B10'] = results.get('total_trucks_deployed', 0)
    ws['A11'] = 'Roads Serviced:'
    ws['B11'] = summary.get('total_roads_serviced', 0)
    ws['A12'] = 'Roads Not Completed:'
    ws['B12'] = summary.get('roads_not_completed', 0)
    ws['A13'] = 'Estimated Completion:'
    ws['B13'] = f"{summary.get('estimated_completion_time_hours', 0)} hours"
    ws['A14'] = 'Total Distance:'
    ws['B14'] = f"{summary.get('total_fleet_distance_miles', 0)} miles"
    
    # Operational Metrics
    ws['D4'] = 'OPERATIONAL METRICS'
    ws['D4'].font = Font(bold=True, size=12)
    ws['D4'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['D4'].font = Font(bold=True, color='FFFFFF', size=12)
    
    ws['D5'] = 'Salt Used:'
    ws['E5'] = f"{summary.get('total_salt_used_tons', 0)} tons"
    ws['D6'] = 'Avg Salt/Truck:'
    ws['E6'] = f"{summary.get('avg_salt_per_truck_tons', 0)} tons"
    ws['D7'] = 'Salt Refills:'
    ws['E7'] = summary.get('total_salt_refills', 0)
    ws['D8'] = 'Driver Breaks:'
    ws['E8'] = summary.get('total_driver_breaks', 0)
    ws['D9'] = 'Shift Compliance:'
    ws['E9'] = f"{summary.get('shift_compliance_percent', 0)}%"
    ws['D10'] = 'Trucks Over Shift:'
    ws['E10'] = summary.get('trucks_over_shift', 0)
    
    # Priority Breakdown
    ws['A16'] = 'PRIORITY DISTRIBUTION'
    ws['A16'].font = Font(bold=True, size=12)
    ws['A16'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A16'].font = Font(bold=True, color='FFFFFF', size=12)
    
    priority = summary.get('priority_breakdown', {})
    ws['A17'] = 'Priority 1 (Emergency):'
    ws['B17'] = priority.get('1', priority.get(1, 0))
    ws['A18'] = 'Priority 2 (High):'
    ws['B18'] = priority.get('2', priority.get(2, 0))
    ws['A19'] = 'Priority 3 (Medium):'
    ws['B19'] = priority.get('3', priority.get(3, 0))
    ws['A20'] = 'Priority 4 (Low):'
    ws['B20'] = priority.get('4', priority.get(4, 0))
    
    # Auto-fit columns
    for col in ['A', 'B', 'D', 'E']:
        ws.column_dimensions[col].width = 25

def create_routes_sheet(ws, results):
    """Detailed truck route information"""
    # Headers
    headers = ['Truck ID', 'Depot', 'Roads', 'Time (hrs)', 'Distance (mi)', 
               'Salt Used (tons)', 'Refills', 'Breaks', 'Shift Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    routes = results.get('routes', [])
    for row, route in enumerate(routes, 2):
        ws.cell(row, 1, route.get('truck_id'))
        ws.cell(row, 2, route.get('depot'))
        ws.cell(row, 3, route.get('roads_serviced'))
        ws.cell(row, 4, round(route.get('total_time_minutes', 0) / 60, 2))
        ws.cell(row, 5, round(route.get('total_distance_miles', 0), 1))
        ws.cell(row, 6, route.get('salt_used_tons', 0))
        ws.cell(row, 7, route.get('refills_needed', 0))
        ws.cell(row, 8, route.get('breaks_taken', 0))
        ws.cell(row, 9, route.get('shift_status', ''))
    
    # Auto-fit columns
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

def create_priority_sheet(ws, results):
    """Priority road analysis with top roads by priority"""
    ws['A1'] = 'PRIORITY ROAD ANALYSIS'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    # Count roads by priority from routes
    priority_roads = {1: [], 2: [], 3: [], 4: []}
    routes = results.get('routes', [])
    
    for route in routes:
        for segment in route.get('route', []):
            if segment.get('segment_type') == 'road':
                priority = segment.get('priority')
                priority_roads[priority].append({
                    'name': segment.get('road_name', 'Unnamed'),
                    'truck': route.get('truck_id'),
                    'time': segment.get('service_time_minutes', 0),
                    'length': segment.get('length_miles', 0),
                    'salt': segment.get('salt_used_lbs', 0)
                })
    
    # Headers
    ws['A3'] = 'Priority'
    ws['B3'] = 'Road Name'
    ws['C3'] = 'Truck ID'
    ws['D3'] = 'Service Time (min)'
    ws['E3'] = 'Length (mi)'
    ws['F3'] = 'Salt (lbs)'
    for col in range(1, 7):
        cell = ws.cell(3, col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    row = 4
    for priority in [1, 2, 3, 4]:
        priority_name = {1: 'EMERGENCY', 2: 'HIGH', 3: 'MEDIUM', 4: 'LOW'}[priority]
        roads = priority_roads[priority][:200]  # Limit to first 200 per priority
        
        for road in roads:
            ws.cell(row, 1, f"P{priority} {priority_name}")
            ws.cell(row, 2, road['name'])
            ws.cell(row, 3, road['truck'])
            ws.cell(row, 4, round(road['time'], 1))
            ws.cell(row, 5, round(road['length'], 2))
            ws.cell(row, 6, round(road['salt'], 0))
            row += 1
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

def create_cost_sheet(ws, results):
    """Cost analysis with assumptions and calculations"""
    ws['A1'] = 'COST ANALYSIS'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:B1')
    
    # Assumptions
    ws['A3'] = 'COST ASSUMPTIONS'
    ws['A3'].font = Font(bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A3:B3')
    
    ws['A4'] = 'Salt cost per ton:'
    ws['B4'] = '$50'
    ws['A5'] = 'Fuel cost per gallon:'
    ws['B5'] = '$3.50'
    ws['A6'] = 'MPG (salt truck):'
    ws['B6'] = '6'
    ws['A7'] = 'Driver wage per hour:'
    ws['B7'] = '$35'
    
    summary = results.get('summary', {})
    total_salt_tons = summary.get('total_salt_used_tons', 0)
    total_distance = summary.get('total_fleet_distance_miles', 0)
    avg_time_hours = summary.get('average_truck_time_hours', 0)
    num_trucks = results.get('total_trucks_deployed', 0)
    
    # Calculations
    salt_cost = total_salt_tons * 50
    fuel_gallons = total_distance / 6
    fuel_cost = fuel_gallons * 3.50
    labor_hours = avg_time_hours * num_trucks
    labor_cost = labor_hours * 35
    total_cost = salt_cost + fuel_cost + labor_cost
    
    ws['A9'] = 'COST BREAKDOWN'
    ws['A9'].font = Font(bold=True, color='FFFFFF')
    ws['A9'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A9:B9')
    
    ws['A10'] = 'Salt cost:'
    ws['B10'] = f"${salt_cost:,.2f}"
    ws['A11'] = 'Fuel cost:'
    ws['B11'] = f"${fuel_cost:,.2f}"
    ws['A12'] = 'Labor cost:'
    ws['B12'] = f"${labor_cost:,.2f}"
    ws['A14'] = 'TOTAL COST:'
    ws['B14'] = f"${total_cost:,.2f}"
    ws['A14'].font = Font(bold=True)
    ws['B14'].font = Font(bold=True)
    
    ws['A16'] = 'PER UNIT COSTS'
    ws['A16'].font = Font(bold=True, color='FFFFFF')
    ws['A16'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A16:B16')
    
    ws['A17'] = 'Cost per mile:'
    ws['B17'] = f"${total_cost/total_distance:,.2f}" if total_distance > 0 else 'N/A'
    ws['A18'] = 'Cost per road:'
    roads_serviced = summary.get('total_roads_serviced', 1)
    ws['B18'] = f"${total_cost/roads_serviced:,.2f}" if roads_serviced > 0 else 'N/A'
    ws['A19'] = 'Cost per truck:'
    ws['B19'] = f"${total_cost/num_trucks:,.2f}" if num_trucks > 0 else 'N/A'
    
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 25

def create_depot_sheet(ws, results):
    """Depot-by-depot performance comparison"""
    ws['A1'] = 'DEPOT PERFORMANCE COMPARISON'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ['Depot', 'Trucks', 'Roads', 'Avg Time (hrs)', 'Total Distance (mi)', 'Salt Used (tons)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Group routes by depot
    depot_stats = {}
    routes = results.get('routes', [])
    
    for route in routes:
        depot = route.get('depot', 'Unknown')
        if depot not in depot_stats:
            depot_stats[depot] = {
                'trucks': 0,
                'roads': 0,
                'total_time': 0,
                'total_distance': 0,
                'total_salt': 0
            }
        
        depot_stats[depot]['trucks'] += 1
        depot_stats[depot]['roads'] += route.get('roads_serviced', 0)
        depot_stats[depot]['total_time'] += route.get('total_time_minutes', 0)
        depot_stats[depot]['total_distance'] += route.get('total_distance_miles', 0)
        depot_stats[depot]['total_salt'] += route.get('salt_used_tons', 0)
    
    # Write data
    row = 4
    for depot, stats in sorted(depot_stats.items()):
        avg_time = stats['total_time'] / stats['trucks'] / 60 if stats['trucks'] > 0 else 0
        ws.cell(row, 1, depot)
        ws.cell(row, 2, stats['trucks'])
        ws.cell(row, 3, stats['roads'])
        ws.cell(row, 4, round(avg_time, 2))
        ws.cell(row, 5, round(stats['total_distance'], 1))
        ws.cell(row, 6, round(stats['total_salt'], 2))
        row += 1
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 30

def create_detailed_routes_sheet(ws, results):
    """Detailed segment-by-segment routes for first 10 trucks"""
    ws['A1'] = 'DETAILED ROUTES (First 10 Trucks)'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')
    
    # Headers
    headers = ['Truck', 'Segment #', 'Type', 'Road Name', 'Priority', 'Time (min)', 'Distance (mi)', 'Salt (lbs)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    row = 4
    routes = results.get('routes', [])[:10]  # First 10 trucks only
    
    for route in routes:
        truck_id = route.get('truck_id')
        for seg_num, segment in enumerate(route.get('route', [])[:50], 1):  # First 50 segments
            ws.cell(row, 1, truck_id)
            ws.cell(row, 2, seg_num)
            ws.cell(row, 3, segment.get('segment_type', ''))
            
            if segment.get('segment_type') == 'road':
                ws.cell(row, 4, segment.get('road_name', ''))
                priority = segment.get('priority', '')
                priority_name = {1: 'P1-EMERG', 2: 'P2-HIGH', 3: 'P3-MED', 4: 'P4-LOW'}.get(priority, '')
                ws.cell(row, 5, priority_name)
                ws.cell(row, 6, round(segment.get('service_time_minutes', 0), 1))
                ws.cell(row, 7, round(segment.get('length_miles', 0), 2))
                ws.cell(row, 8, round(segment.get('salt_used_lbs', 0), 0))
            else:
                ws.cell(row, 4, segment.get('description', ''))
                ws.cell(row, 6, round(segment.get('time_minutes', 0), 1))
            
            row += 1
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20

if __name__ == "__main__":
    print("="*70)
    print("PITTSBURGH SALT TRUCK - EXCEL REPORT GENERATOR")
    print("="*70)
    
    try:
        create_excel_report()
        print("\n✓ Excel report generation complete!")
        print("\nThe Excel file contains:")
        print("  • Executive Summary - Key metrics and overview")
        print("  • All Truck Routes - Complete fleet data")
        print("  • Priority Analysis - Roads by priority level")
        print("  • Cost Analysis - Estimated operational costs")
        print("  • Depot Performance - Comparison across depots")
        print("  • Detailed Routes - Segment-by-segment breakdown")
    except FileNotFoundError:
        print("\nERROR: optimization_results.json not found!")
        print("Please run PGH_Salt_Trucks_FINAL.py first to generate the JSON file.")
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
